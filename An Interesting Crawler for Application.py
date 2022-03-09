import requests
import openpyxl
import time
import json

def gettalent(flag):
    try:
        url = "https://api.gotokeep.com/social/v3/verified/profile/label/5d1adc9e3f3d41d7991c19c0?limit=20&page=" + str(flag)
        timestamp = str(int(float(time.time())*1000))
        headers = {
            # Please add your personal header there
        }
        # print(type(json.dumps(data)))
        resp = requests.get(url,headers=headers)
        # print(resp.text)
        users_list = json.loads(resp.text).get("data").get("profiles")
        normal = []
        user = []
        userprofile = []

        num = len(users_list)
        # print(users_list)
        for i in range(len(users_list)):
            useri, userprofilei, normali= get_a_person_keys(users_list[i])
            user += useri
            userprofile += userprofilei
            normal += normali

            user = uniquee(user)
            userprofile = uniquee(userprofile)
            normal = uniquee(normal)

        return user, userprofile, normal,num                       #### Find all feature and create a union                                 
    except:
        time.sleep(1)
        user, userprofile, normal, num = gettalent(flag)
        return user, userprofile, normal, num
    
def get_a_person_keys(user_dist):
    # print(user_dist)
    normal = []
    user = []
    userprofile = []

    entries = []

    for keys in user_dist.keys():
        # print(keys)
        items = user_dist[keys]
        # if keys == "labels":
        #     if len(items)>0:
        #         for i in range(len(items)):
        #             labels.append(items[i])
        if keys == "entries":
            if len(items)>0:
                for i in range(len(items)):
                    entries += get_dict_info(items[i])  
        elif keys == "user":
            user += get_dict_info(items)
        elif keys == "userProfile":
            userprofile += get_dict_info(items)           
        else:
            normal.append(keys)

    return user, userprofile, normal

def get_dict_info(a_dict):
        sub_keys = []

        for keys in a_dict.keys():
            sub_keys.append(keys)
        
        return sub_keys

def save(user, userprofile, normal):
        # wb = openpyxl.load_workbook('talent.xlsx')
        # ws = wb[wb.sheetnames[0]]
        # for i in range(len(data)):
        #     ws.append(data[i])
        # wb.save('talent.xlsx')
        wb = openpyxl.load_workbook('keys244.xlsx')
        ws = wb[wb.sheetnames[0]]

        ws.append(user)
        ws.append(userprofile)
        ws.append(normal)
        entries = ["photo", "type", "vedio", "_id"]*3
        ws.append(entries)

        wb.save('keys244.xlsx')

def uniquee(a_list):
    l1 = []
    for i in a_list:
        if i not in l1:
            l1.append(i)
    return l1  

if __name__ == '__main__':
        flag = 1
        user = []
        userprofile = []
        normal = []

        while flag:
            useri, userprofilei, normali,numi = gettalent(flag)
            user += useri
            userprofile += userprofilei
            normal += normali



            print("Already have{} data".format(flag*20))

            if flag == 14:
                break

            flag += 1
            time.sleep(0.1)

        user = uniquee(user)
        userprofile = uniquee(userprofile)
        normal = uniquee(normal)


        save(user, userprofile, normal)

